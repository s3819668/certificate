from openpyxl import load_workbook
import datetime
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

wb = load_workbook("Certificate.xlsx")
for sheet in wb:
    colE=sheet['E']
    colF=sheet['F']
    orange = PatternFill(start_color='ff8000', end_color='ff8000', fill_type='solid')
    yellow = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
    sheet.conditional_formatting.add('E1:E99999', FormulaRule(formula=['OR($E1="no peer certificate available",$E1="unable to load certificate",$E1="return empty")'], stopIfTrue=False, fill=orange))
    sheet.conditional_formatting.add('F1:F99999', FormulaRule(formula=['$F1="N/A"'], stopIfTrue=False, fill=orange))
    sheet.conditional_formatting.add('F1:F99999', FormulaRule(formula=['AND($F1<Sheet!$G$2,$F1<>"")'], stopIfTrue=False, fill=yellow))
wb.save("Certificate.xlsx")#存檔