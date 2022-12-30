from datetime import datetime
from openpyxl import Workbook
import os
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule


#日期計算
class machine():
    def __init__(self, RC, site, name, expire_day):
        self.RC = RC
        self.ware = 'SW'  # pre-define SW if have HW then assign HW
        if 'HW' in site:
            self.ware = 'HW'
        self.site = site
        self.name = name
        self.expire_day = expire_day
        self.remain = 'N/A'  # 預設N/A天
        if expire_day != 'return empty' and expire_day != 'no peer certificate available' and expire_day != 'unable to load certificate':  # 非N/A日期相減
            M = {'Jan': '1', 'Feb': '2', 'Mar': '3', 'Apr': '4', 'May': '5', 'Jun': '6', 'Jul': '7', 'Aug': '8',
                 'Sep': '9', 'Oct': '10', 'Nov': '11', 'Dec': '12'}
            notAfter = self.expire_day.replace(self.expire_day[0:3], M[self.expire_day[0:3]])  # 月份簡稱換成月份數字
            if "\n" in notAfter:
                notAfter = notAfter[:-5]  # 去除GMT\n 可以替換成 not.remove("GMT")
            else:
                notAfter = notAfter[:-4]  # 去除GMT 可以替換成 not.remove("GMT")因為最後一行貼上未必有\n 判斷避免最後日期變成empty
            notAfter = datetime.strptime(notAfter, '%m %d %H:%M:%S %Y')  # 正規化日期
            self.expire_day = notAfter.date()
            self.remain = (notAfter - datetime.now()).days  # 相減


def get_data(path):
    all_vm = []  # 存放VM物件的容器
    # slide window algorithm 以\n做為切割提取物件
    for file in os.listdir(path):  # 走訪dir資料夾
        # slide window algorithm指標
        l = 0
        r = 0
        with open(path+'/' + file, 'r', encoding='UTF-8') as f:  # 讀檔
            read_data = [i for i in f.readlines()]
            for i, data in enumerate(read_data):
                if data == "\n" or data == " \n" or i == len(read_data) - 1:  # 處理條件 遇到空白行或者最後一行做分割段行
                    r = i  # 遇到處理條件 右邊指標=當前指標i  擷取l:r即當前VM
                    # 處理物件沒有notAfter 造成日期格式format except問題
                    try:
                        print([j for j in read_data[l:r + 1] if "notAfter" in j][-1][9:])
                        # 建立物件 RC用檔名'.'去做切割 再去除最後一項拿掉txt的副檔名 ,用l~r+1 取得txt的第l行到第r行, 找到l~r+1 中的notAfter行 9:拿掉netAfter=
                        all_vm.append(machine(file.split(" ")[0] if " " in file else file.split('.')[0],
                                              ''.join(file.split(".")[:-1]), read_data[l:r + 1][0],
                                              [j for j in read_data[l:r + 1] if "notAfter" in j][-1][9:]))
                    except Exception as e:
                        print(e)
                        if any([True for i in read_data[l:r + 1] if 'no peer certificate available' in i]):
                            all_vm.append(machine(file.split(" ")[0] if " " in file else file.split('.')[0],
                                                  ''.join(file.split(".")[:-1]), read_data[l:r + 1][0],
                                                  'no peer certificate available'))
                        elif any([True for i in read_data[l:r + 1] if 'unable to load certificate' in i]):
                            all_vm.append(machine(file.split(" ")[0] if " " in file else file.split('.')[0],
                                                  ''.join(file.split(".")[:-1]), read_data[l:r + 1][0],
                                                  'unable to load certificate'))
                        else:
                            # 同上只是少傳一個參數 利用物件預設參數expire_day=N/A 設定expire_day為N/A 順便不觸發if 使remain=預設N/A天
                            all_vm.append(machine(file.split(" ")[0] if " " in file else file.split('.')[0],
                                                  ''.join(file.split(".")[:-1]), read_data[l:r + 1][0], 'return empty'))
                    l = r + 1  # 更新l指標準備做為下一個VM的開頭
                else:
                    if "\n" in read_data[i]:
                        read_data[i] = read_data[i].rstrip()  # 去除\n符號,必須寫在else不然欲處理的\n判斷失敗
    return all_vm


def sheet_classification(wb,sheet_name, classification_data):
    def sheet_init():
        # 標題title
        sheet.cell(1, 1).value = 'RC'
        sheet.cell(1, 2).value = 'SW/HW'
        sheet.cell(1, 3).value = 'site'
        sheet.cell(1, 4).value = 'VM_name'
        sheet.cell(1, 5).value = 'expire_day'
        sheet.cell(1, 6).value = 'remain_days'

    def sheet_fill_value(vms):
        for i in range(len(vms)):  # 走訪物件excel從1,1開始 所以i+1 ,1234 ,預留第一行給標題所以i在加1=>i+2
            sheet.cell(i + 2, 1).value = vms[i].RC
            sheet.cell(i + 2, 2).value = vms[i].ware
            sheet.cell(i + 2, 3).value = vms[i].site
            sheet.cell(i + 2, 4).value = vms[i].name
            sheet.cell(i + 2, 5).value = vms[i].expire_day
            sheet.cell(i + 2, 6).value = '=IF(E' + str(i + 2) + '="","",IFERROR(E' + str(i + 2) + '-TODAY(),"N/A"))'

    def sheet_fill_color_rule():
        orange = PatternFill(start_color='ff8000', end_color='ff8000', fill_type='solid')
        yellow = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
        sheet.conditional_formatting.add('E1:E20000', FormulaRule(
            formula=['OR($E1="no peer certificate available",$E1="unable to load certificate",$E1="return empty")'],
            stopIfTrue=False, fill=orange))
        sheet.conditional_formatting.add('F1:F20000', FormulaRule(formula=['$F1="N/A"'], stopIfTrue=False, fill=orange))
        sheet.conditional_formatting.add('F1:F20000',
                                         FormulaRule(formula=['AND($F1<Sheet!$G$2,$F1<>"")'], stopIfTrue=False,
                                                     fill=yellow))
    try:
        sheet = wb[sheet_name]  # 指定工作簿
        sheet.cell(1, 7).value = 'mark yellow if remain_days less than this days'
        sheet.cell(2, 7).value = 365
    except:
        wb.create_sheet(sheet_name)
        sheet = wb[sheet_name]  # 指定工作簿

    sheet_init()
    sheet_fill_value(classification_data)
    sheet_fill_color_rule()


if __name__=="__main__":
    all_vm=get_data('certificate')
    # 寫入excel程式
    wb = Workbook()  # 建立excel物件
    sheet_classification(wb,'Sheet',all_vm)
    classification=["emc","pnp","cls","vm","Switch","oa","iLO","esxi","vcenter","vc","ave","hpsim","vdp","other"]
    classification={i:[] for i in classification}

    for i in all_vm:
        if "spa" in i.name or "spb" in i.name:
            classification["emc"].append(i)
        elif "pnp" in i.name:
            classification["pnp"].append(i)
        elif "cls" in i.name:
            classification["cls"].append(i)
        elif "vm" in i.name:
            classification["vm"].append(i)
        elif "sw" in i.name:
            classification["Switch"].append(i)
        elif "oa" in i.name:
            classification["oa"].append(i)
        elif "iLO" in i.name:
            classification["iLO"].append(i)
        elif "esxi" in i.name:
            classification["esxi"].append(i)
        elif "vcenter" in i.name:
            classification["vcenter"].append(i)
        elif "vc" in i.name:
            classification["vc"].append(i)
        elif "ave" in i.name or "bve" in i.name:
            classification["ave"].append(i)
        elif "hpsim" in i.name:
            classification["hpsim"].append(i)
        elif "vdp" in i.name:
            classification["vdp"].append(i)
        else:
            classification["other"].append(i)
    for i in classification:
        sheet_classification(wb, i, classification[i])

    wb.save("Certificate.xlsx")  # 存檔